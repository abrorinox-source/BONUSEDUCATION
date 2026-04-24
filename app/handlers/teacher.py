"""
Teacher handlers
All teacher-related functionality
"""

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramBadRequest
from aiogram.utils.keyboard import InlineKeyboardBuilder
from app.database import db
from app.sheets_manager import sheets_manager
from app import keyboards
from app.states import AddPointsStates, SubtractPointsStates, BroadcastStates, EditRulesStates, GroupStates, SettingsStates
from app import config
from datetime import datetime

router = Router()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS FOR ERROR HANDLING
# ═══════════════════════════════════════════════════════════════════════════════

async def safe_edit_message(callback: CallbackQuery, text: str, reply_markup=None):
    """
    Safely edit message with proper error handling
    Handles 'message is not modified' errors gracefully
    """
    try:
        await callback.message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e):
            # Message content is the same, just answer the callback
            await safe_answer_callback(callback)
        else:
            raise


async def safe_answer_callback(callback: CallbackQuery, text: str = None, show_alert: bool = False):
    """
    Safely answer callback query with proper error handling
    Handles 'query is too old' errors gracefully
    """
    try:
        await callback.answer(text, show_alert=show_alert)
    except TelegramBadRequest as e:
        if "query is too old" in str(e):
            print(f"Callback query expired, ignoring: {callback.data}")
        else:
            raise


def format_transfer_limits_text() -> str:
    """Build transfer limits settings text."""
    settings = db.get_transfer_limit_settings()
    return (
        "🚦 <b>Transfer Limits</b>\n\n"
        f"📅 <b>Daily transfer count:</b> {settings['daily_transfer_count_limit'] or 'Unlimited'}\n"
        f"🗓️ <b>Weekly transfer count:</b> {settings['weekly_transfer_count_limit'] or 'Unlimited'}\n"
        f"💸 <b>Daily transfer points:</b> {settings['daily_transfer_points_limit'] or 'Unlimited'}\n"
        f"📦 <b>Weekly transfer points:</b> {settings['weekly_transfer_points_limit'] or 'Unlimited'}\n\n"
        "<i>0 means unlimited.</i>\n"
        "<i>These limits are counted per sender only.</i>\n"
        "<i>Only users with real Telegram IDs are tracked.</i>"
    )


def format_student_transfer_limits_text(user_id: str, student_name: str) -> str:
    """Build per-student transfer limit text."""
    effective = db.get_effective_transfer_limits(user_id)
    override = db.get_transfer_limit_override(user_id)
    return (
        f"🚦 <b>Transfer Limits for {student_name}</b>\n\n"
        f"📅 <b>Daily transfer count:</b> {effective['daily_transfer_count_limit'] or 'Unlimited'}"
        f" (override: {override['daily_transfer_count_limit'] or 'Global'})\n"
        f"🗓️ <b>Weekly transfer count:</b> {effective['weekly_transfer_count_limit'] or 'Unlimited'}"
        f" (override: {override['weekly_transfer_count_limit'] or 'Global'})\n"
        f"💸 <b>Daily transfer points:</b> {effective['daily_transfer_points_limit'] or 'Unlimited'}"
        f" (override: {override['daily_transfer_points_limit'] or 'Global'})\n"
        f"📦 <b>Weekly transfer points:</b> {effective['weekly_transfer_points_limit'] or 'Unlimited'}"
        f" (override: {override['weekly_transfer_points_limit'] or 'Global'})\n\n"
        "<i>Override value 0 means use the global setting.</i>"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN MENU HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.message(F.text.contains("Refresh Groups"))
async def refresh_groups_menu(message: Message):
    """Refresh groups cache from Google Sheets"""
    teacher_id = str(message.from_user.id)

    # Show loading message
    loading_msg = await message.answer("🔄 Refreshing groups from Google Sheets...")

    # Force refresh from Google Sheets
    groups = db.get_teacher_groups(teacher_id, force_refresh=True)

    # Delete loading message
    await loading_msg.delete()

    # Show results
    if groups:
        text = "✅ Groups refreshed successfully!\n\n"
        text += f"📊 You have {len(groups)} group(s):\n"
        for group in groups:
            student_count = group.get('student_count', 0)
            text += f"  • {group['name']} - {student_count} student(s)\n"
        text += f"\n🕐 Last updated: just now"
    else:
        text = "❌ No groups found.\n\nCreate your first group in:\nSettings → Manage Groups"

    await message.answer(text, reply_markup=keyboards.get_teacher_keyboard())


@router.message(F.text.contains("Force Sync"))
async def force_sync(message: Message):
    """Sheets-only mode: manual sync is disabled."""
    await message.answer(
        "🔄 Force Sync is disabled in Sheets-only mode.\n\nUse Refresh Groups or edit the Sheet tabs directly."
    )


@router.message(F.text.contains("Rating"))
async def show_rating_all(message: Message, user: dict = None):
    """Show rating - auto for students, selection for teachers"""
    user_id = str(message.from_user.id)

    # Get user data if not provided
    if not user:
        user = db.get_user(user_id)

    # Check role
    if user and user.get('role') == 'teacher':
        # Teacher: show group selection
        teacher_id = user_id
        groups = db.get_teacher_groups(teacher_id)

        if not groups:
            await message.answer(
                "❌ You don't have any groups yet.\n"
                "Create a group in Google Sheets first."
            )
            return

        await message.answer(
            "🏆 SELECT GROUP TO VIEW RATING\n\n"
            "Choose a group:",
            reply_markup=keyboards.get_group_selection_keyboard(groups, "rating")
        )
    else:
        # Student: show rating directly (no selection)
        group_id = user.get('group_id') if user else None
        if not group_id:
            await message.answer("❌ You are not assigned to any group yet.")
            return

        # Get ranking directly
        ranking = db.get_ranking(group_id=group_id)
        group = db.get_group(group_id)
        group_name = group.get('name', group_id) if group else group_id

        if not ranking:
            await message.answer(f"📊 No students found in {group_name}.")
            return

        page = 0
        total_pages = max(1, (len(ranking) + PAGE_SIZE_RANKING - 1) // PAGE_SIZE_RANKING)
        text = build_ranking_text_teacher(ranking, group_name, page, PAGE_SIZE_RANKING, highlight_user_id=user_id)

        await message.answer(
            text,
            parse_mode="Markdown",
            reply_markup=keyboards.get_ranking_keyboard("student", page, total_pages, group_id)
        )


@router.message(F.text.contains("Students"))
async def show_students(message: Message):
    """Show group selection for students list"""
    teacher_id = str(message.from_user.id)
    groups = db.get_teacher_groups(teacher_id)

    if not groups:
        await message.answer("❌ No groups found. Create a group first in Settings → Manage Groups.")
        return

    # Build keyboard with all groups + "All Students" option
    builder = InlineKeyboardBuilder()

    # Add individual groups
    for group in groups:
        student_count = group.get('student_count', 0)
        builder.button(
            text=f"📁 {group['name']} ({student_count} students)",
            callback_data=f"students:group:{group['group_id']}"
        )

    # Add "All Students" option
    builder.button(
        text="👥 All Students",
        callback_data="students:all"
    )

    builder.adjust(1)

    await message.answer(
        "👤 SELECT GROUP TO VIEW STUDENTS\n\n"
        "Choose a group:",
        reply_markup=builder.as_markup()
    )


@router.message(F.text.contains("Settings"))
async def show_settings(message: Message):
    """Show settings menu"""
    await message.answer(
        "⚙️ BOT SETTINGS\n"
        "Select an option:",
        reply_markup=keyboards.get_settings_keyboard()
    )


# ═══════════════════════════════════════════════════════════════════════════════
# STUDENT MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data.startswith("student_detail:"))
async def student_detail(callback: CallbackQuery):
    """Show student details"""
    parts = callback.data.split(":")
    user_id = parts[1]
    scope = parts[2] if len(parts) > 2 else "all"
    student = db.get_user(user_id)

    if not student:
        await callback.answer("❌ Student not found!", show_alert=True)
        return

    student_id_label = "Manual entry" if student.get('is_manual') else student.get('user_id', 'N/A')

    text = (
        f"👤 <b>Student Details</b>\n\n"
        f"📝 <b>Name:</b> {student['full_name']}\n"
        f"📞 <b>Phone:</b> {student.get('phone', 'N/A')}\n"
        f"🔗 <b>Username:</b> @{student.get('username', 'N/A')}\n"
        f"💎 <b>Points:</b> {student['points']}\n"
        f"📌 <b>Status:</b> {student['status']}\n"
        f"🆔 <b>ID:</b> {student_id_label}\n\n"
        f"<i>Choose an action below.</i>"
    )

    await callback.message.edit_text(
        text,
        reply_markup=keyboards.get_student_detail_keyboard(
            user_id,
            back_target="students:list" if scope == "all" else f"students:group:{scope}"
        )
    )
    await callback.answer()


@router.callback_query(F.data.startswith("stl:"))
async def student_transfer_limits(callback: CallbackQuery, state: FSMContext):
    """Manage per-student transfer limit overrides."""
    parts = callback.data.split(":")
    if len(parts) < 2:
        await safe_answer_callback(callback, "Invalid transfer limits action", show_alert=True)
        return

    if parts[1] in {"set", "reset"}:
        action = parts[1]
        user_id = parts[2] if len(parts) > 2 else None
    else:
        action = "view"
        user_id = parts[1]

    if not user_id:
        await safe_answer_callback(callback, "Student not found", show_alert=True)
        return

    key_map = {
        'dc': 'daily_transfer_count_limit',
        'wc': 'weekly_transfer_count_limit',
        'dp': 'daily_transfer_points_limit',
        'wp': 'weekly_transfer_points_limit',
    }
    labels = {
        'daily_transfer_count_limit': 'daily transfer count limit',
        'weekly_transfer_count_limit': 'weekly transfer count limit',
        'daily_transfer_points_limit': 'daily transfer points limit',
        'weekly_transfer_points_limit': 'weekly transfer points limit',
    }

    if action == "set":
        if len(parts) < 4:
            await safe_answer_callback(callback, "Invalid transfer limits action", show_alert=True)
            return
        setting_key = key_map.get(parts[3])
        if not setting_key:
            await safe_answer_callback(callback, "Invalid transfer limit key", show_alert=True)
            return
        student = db.get_user(user_id)
        if not student:
            await safe_answer_callback(callback, "Student not found", show_alert=True)
            return
        if not str(user_id).isdigit():
            await safe_answer_callback(callback, "This student has no Telegram ID", show_alert=True)
            return

        await state.set_state(SettingsStates.waiting_for_transfer_limit)
        await state.update_data(
            transfer_limit_key=setting_key,
            transfer_limit_label=labels[setting_key],
            transfer_limit_target_user_id=user_id,
            transfer_limit_target_name=student['full_name'],
        )
        await callback.message.edit_text(
            f"Enter new value for {student['full_name']} - {labels[setting_key]}.\n\n"
            f"Send 0 to use the global setting.\n"
            f"Only whole numbers are allowed.\n\n"
            f"Send /cancel to abort."
        )
        await safe_answer_callback(callback)
        return

    if action == "reset":
        student = db.get_user(user_id)
        if not student:
            await safe_answer_callback(callback, "Student not found", show_alert=True)
            return
        if not db.reset_transfer_limit_override(user_id):
            await safe_answer_callback(callback, "Reset failed", show_alert=True)
            return
        await safe_edit_message(
            callback,
            format_student_transfer_limits_text(user_id, student['full_name']),
            reply_markup=keyboards.get_student_transfer_limits_keyboard(
                user_id, db.get_transfer_limit_override(user_id)
            )
        )
        await safe_answer_callback(callback, "Override reset")
        return

    student = db.get_user(user_id)
    if not student:
        await safe_answer_callback(callback, "Student not found", show_alert=True)
        return
    if not str(user_id).isdigit():
        await safe_answer_callback(callback, "This student has no Telegram ID", show_alert=True)
        return

    await safe_edit_message(
        callback,
        format_student_transfer_limits_text(user_id, student['full_name']),
        reply_markup=keyboards.get_student_transfer_limits_keyboard(
            user_id, db.get_transfer_limit_override(user_id)
        )
    )
    await safe_answer_callback(callback)


@router.callback_query(F.data.startswith("add_points:"))
async def add_points_start(callback: CallbackQuery, state: FSMContext):
    """Start add points flow"""
    user_id = callback.data.split(":")[1]
    student = db.get_user(user_id)

    if not student:
        await callback.answer("❌ Student not found!", show_alert=True)
        return

    await state.update_data(target_user_id=user_id, target_user_name=student['full_name'])
    await state.set_state(AddPointsStates.waiting_for_amount)

    await callback.message.answer(
        f"➕ ADD POINTS\n"
        f"Student: {student['full_name']}\n\n"
        f"Enter amount to add (1-10000):"
    )
    await callback.answer()


@router.message(AddPointsStates.waiting_for_amount)
async def add_points_amount(message: Message, state: FSMContext):
    """Process add points amount"""
    try:
        amount = int(message.text.strip())

        if amount <= 0 or amount > 10000:
            await message.answer("❌ Amount must be between 1 and 10000. Try again:")
            return

        data = await state.get_data()
        student = db.get_user(data['target_user_id'])

        text = (
            f"⚠️ CONFIRMATION\n"
            f"Student: {data['target_user_name']}\n"
            f"Action: Add Points\n"
            f"Amount: {amount} pts\n"
            f"Current Balance: {student['points']} pts\n"
            f"New Balance: {student['points'] + amount} pts\n\n"
            f"Confirm this action?"
        )

        await state.update_data(amount=amount)

        await message.answer(
            text,
            reply_markup=keyboards.get_confirmation_keyboard("add_points", data['target_user_id'])
        )

    except ValueError:
        await message.answer("❌ Please enter a valid number:")


@router.callback_query(F.data.startswith("confirm:add_points:"))
async def confirm_add_points(callback: CallbackQuery, state: FSMContext):
    """Confirm and execute add points"""
    user_id = callback.data.split(":")[2]
    data = await state.get_data()
    amount = data['amount']
    teacher_id = str(callback.from_user.id)

    # Add points (atomic)
    result = db.add_points(user_id, amount)

    if result['success']:
        # Log transaction
        db.log_add_points(
            teacher_id=teacher_id,
            student_id=user_id,
            amount=amount,
            student_name=data['target_user_name'],
            old_balance=result['new_balance'] - amount,
            new_balance=result['new_balance']
        )

        await callback.message.edit_text(
            f"✅ Points Added!\n"
            f"Student: {data['target_user_name']}\n"
            f"Amount: +{amount} pts\n"
            f"New Balance: {result['new_balance']} pts"
        )

        # Notify student
        try:
            await callback.bot.send_message(
                chat_id=user_id,
                text=f"💰 Teacher added {amount} pts to your account!\n"
                     f"New balance: {result['new_balance']} pts"
            )
        except:
            pass
    else:
        await callback.message.edit_text(f"❌ Error: {result['error']}")

    await state.clear()
    await callback.answer()


@router.callback_query(F.data.startswith("subtract_points:"))
async def subtract_points_start(callback: CallbackQuery, state: FSMContext):
    """Start subtract points flow"""
    user_id = callback.data.split(":")[1]
    student = db.get_user(user_id)

    if not student:
        await callback.answer("❌ Student not found!", show_alert=True)
        return

    await state.update_data(
        target_user_id=user_id,
        target_user_name=student['full_name'],
        current_balance=student['points']
    )
    await state.set_state(SubtractPointsStates.waiting_for_amount)

    await callback.message.answer(
        f"➖ SUBTRACT POINTS\n"
        f"Student: {student['full_name']}\n"
        f"Current Balance: {student['points']} pts\n\n"
        f"Enter amount to subtract:"
    )
    await callback.answer()


@router.message(SubtractPointsStates.waiting_for_amount)
async def subtract_points_amount(message: Message, state: FSMContext):
    """Process subtract points amount"""
    try:
        amount = int(message.text.strip())
        data = await state.get_data()

        if amount <= 0:
            await message.answer("❌ Amount must be positive. Try again:")
            return

        if amount > data['current_balance']:
            await message.answer(
                f"❌ Amount exceeds current balance ({data['current_balance']} pts).\n"
                f"Try again:"
            )
            return

        student = db.get_user(data['target_user_id'])

        text = (
            f"⚠️ CONFIRMATION\n"
            f"Student: {data['target_user_name']}\n"
            f"Action: Subtract Points\n"
            f"Amount: {amount} pts\n"
            f"Current Balance: {student['points']} pts\n"
            f"New Balance: {student['points'] - amount} pts\n\n"
            f"Confirm this action?"
        )

        await state.update_data(amount=amount)

        await message.answer(
            text,
            reply_markup=keyboards.get_confirmation_keyboard("subtract_points", data['target_user_id'])
        )

    except ValueError:
        await message.answer("❌ Please enter a valid number:")


@router.callback_query(F.data.startswith("confirm:subtract_points:"))
async def confirm_subtract_points(callback: CallbackQuery, state: FSMContext):
    """Confirm and execute subtract points"""
    user_id = callback.data.split(":")[2]
    data = await state.get_data()
    amount = data['amount']
    teacher_id = str(callback.from_user.id)

    # Subtract points (atomic)
    result = db.subtract_points(user_id, amount)

    if result['success']:
        # Log transaction
        db.log_subtract_points(
            teacher_id=teacher_id,
            student_id=user_id,
            amount=amount,
            student_name=data['target_user_name'],
            old_balance=result['new_balance'] + amount,
            new_balance=result['new_balance']
        )

        await callback.message.edit_text(
            f"✅ Points Subtracted!\n"
            f"Student: {data['target_user_name']}\n"
            f"Amount: -{amount} pts\n"
            f"New Balance: {result['new_balance']} pts"
        )

        # Notify student
        try:
            await callback.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ Teacher removed {amount} pts from your account.\n"
                     f"New balance: {result['new_balance']} pts"
            )
        except:
            pass
    else:
        await callback.message.edit_text(f"❌ Error: {result['error']}")

    await state.clear()
    await callback.answer()


@router.callback_query(F.data.startswith("delete_student:"))
async def delete_student_confirm(callback: CallbackQuery):
    """Show delete confirmation"""
    user_id = callback.data.split(":")[1]
    student = db.get_user(user_id)

    if not student:
        await callback.answer("❌ Student not found!", show_alert=True)
        return

    text = (
        f"⚠️ WARNING\n"
        f"Are you sure you want to delete this student?\n\n"
        f"Name: {student['full_name']}\n"
        f"Points: {student['points']} (will be lost)\n\n"
        f"This action CANNOT be undone!"
    )

    await callback.message.edit_text(
        text,
        reply_markup=keyboards.get_confirmation_keyboard("delete_student", user_id)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("confirm:delete_student:"))
async def confirm_delete_student(callback: CallbackQuery):
    """Execute student deletion"""
    user_id = callback.data.split(":")[2]
    student = db.get_user(user_id)

    if not student:
        await callback.answer("❌ Student not found!", show_alert=True)
        return

    # Delete from Sheets
    db.delete_user(user_id)


    # Notify student
    try:
        await callback.bot.send_message(
            chat_id=user_id,
            text=config.MESSAGES['user_deleted']
        )
    except:
        pass

    await callback.message.edit_text(
        f"✅ Student deleted successfully.\n"
        f"Name: {student['full_name']}"
    )
    await callback.answer("✅ Student deleted!")


# ═══════════════════════════════════════════════════════════════════════════════
# CANCEL HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data.startswith("cancel:"))
async def cancel_action(callback: CallbackQuery, state: FSMContext):
    """Cancel current action"""
    await state.clear()
    await callback.message.edit_text("❌ Action cancelled.")
    await callback.answer()


@router.callback_query(F.data == "students:all")
async def show_all_students(callback: CallbackQuery):
    """Show all students (no group filter)"""
    students = db.get_all_users(role='student', status='active')

    if not students:
        await callback.message.edit_text("👤 No active students found.")
        await callback.answer()
        return

    await safe_edit_message(
        callback,
        f"👤 ALL STUDENTS ({len(students)})\n"
        f"Select a student to manage:",
        reply_markup=keyboards.get_students_list_keyboard(students, page=0, scope="all")
    )
    await callback.answer()


@router.callback_query(F.data.startswith("students:group:"))
async def show_group_students(callback: CallbackQuery):
    """Show students from a specific group"""
    group_id = callback.data.split(":")[2]

    # Get group info
    group = db.get_group(group_id)
    if not group:
        await callback.answer("❌ Group not found!", show_alert=True)
        return

    # Get students in this group
    students = db.get_all_users(role='student', status='active', group_id=group_id)

    if not students:
        await safe_edit_message(
            callback,
            f"👤 No students found in {group['name']}.",
            reply_markup=keyboards.get_back_keyboard("teacher:menu")
        )
        await callback.answer()
        return

    await safe_edit_message(
        callback,
        f"👤 {group['name'].upper()} STUDENTS ({len(students)})\n"
        f"Select a student to manage:",
        reply_markup=keyboards.get_students_list_keyboard(students, page=0, scope=group_id)
    )
    await callback.answer()


@router.callback_query(F.data == "students:list")
async def back_to_students_list(callback: CallbackQuery):
    """Return to students list"""
    students = db.get_all_users(role='student', status='active')

    await safe_edit_message(
        callback,
        f"👤 STUDENTS ({len(students)})\n"
        f"Select a student to manage:",
        reply_markup=keyboards.get_students_list_keyboard(students, page=0, scope="all")
    )
    await callback.answer()


@router.callback_query(F.data.startswith("students_page:"))
async def students_page_handler(callback: CallbackQuery):
    """Handle students list pagination"""
    parts = callback.data.split(":")
    scope = parts[1] if len(parts) > 2 else "all"
    page = int(parts[2]) if len(parts) > 2 else int(parts[1])

    if scope == "all":
        students = db.get_all_users(role='student', status='active')
        title = f"👤 ALL STUDENTS ({len(students)})"
    else:
        group = db.get_group(scope)
        if not group:
            await callback.answer("❌ Group not found!", show_alert=True)
            return
        students = db.get_all_users(role='student', status='active', group_id=scope)
        title = f"👤 {group['name'].upper()} STUDENTS ({len(students)})"

    if not students:
        await callback.answer("👤 No students found!", show_alert=True)
        return

    total_pages = max(1, (len(students) + PAGE_SIZE_STUDENTS - 1) // PAGE_SIZE_STUDENTS)
    page = max(0, min(page, total_pages - 1))

    await safe_edit_message(
        callback,
        f"{title}\n"
        f"Sahifa {page + 1}/{total_pages}\n"
        f"Select a student to manage:",
        reply_markup=keyboards.get_students_list_keyboard(students, page=page, scope=scope)
    )
    await callback.answer()


@router.callback_query(F.data.startswith("ranking_page:"))
async def ranking_page_teacher(callback: CallbackQuery):
    """Handle ranking pagination for teacher"""
    # Format: ranking_page:{role}:{group_id}:{page}
    parts = callback.data.split(":")
    role = parts[1]
    group_id = parts[2]
    page = int(parts[3])

    group = db.get_group(group_id)
    group_name = group.get('name', 'Unknown Group') if group else 'Unknown Group'

    ranking = db.get_ranking(group_id=group_id)
    total_pages = max(1, (len(ranking) + PAGE_SIZE_RANKING - 1) // PAGE_SIZE_RANKING)
    page = max(0, min(page, total_pages - 1))

    text = build_ranking_text_teacher(ranking, group_name, page, PAGE_SIZE_RANKING)

    await callback.message.edit_text(
        text,
        reply_markup=keyboards.get_ranking_keyboard(role, page, total_pages, group_id)
    )
    await callback.answer()


PAGE_SIZE_RANKING = 20
PAGE_SIZE_STUDENTS = 10


def build_ranking_text_teacher(ranking: list, group_name: str, page: int, page_size: int, highlight_user_id: str = "") -> str:
    """Build ranking text for teacher (and student fallback) view"""
    total = len(ranking)
    total_pages = max(1, (total + page_size - 1) // page_size)
    start = page * page_size
    end = start + page_size
    page_ranking = ranking[start:end]

    text = f"🏆 {group_name.upper()} RANKING\n"
    if total_pages > 1:
        text += f"Sahifa {page + 1}/{total_pages}\n"
    text += "\n"

    for i, student in enumerate(page_ranking, start + 1):
        emoji = "👑" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        name = student['full_name']
        if highlight_user_id and student.get('user_id') == highlight_user_id:
            name = f"**{name}**"
        text += f"{emoji} {name} - {student['points']} pts\n"

    text += f"\nJami o'quvchilar: {total}"
    return text


@router.callback_query(F.data.startswith("rating:group:"))
async def show_group_rating(callback: CallbackQuery):
    """Show ranking for selected group"""
    group_id = callback.data.split(":")[2]

    # Get group info
    group = db.get_group(group_id)
    if not group:
        await callback.answer("❌ Group not found!", show_alert=True)
        return

    # Get ranking for this group
    ranking = db.get_ranking(group_id=group_id)

    if not ranking:
        await callback.message.edit_text(
            f"📊 No students found in {group['name']}.",
            reply_markup=keyboards.get_back_keyboard("teacher:menu")
        )
        return

    page = 0
    total_pages = max(1, (len(ranking) + PAGE_SIZE_RANKING - 1) // PAGE_SIZE_RANKING)
    text = build_ranking_text_teacher(ranking, group['name'], page, PAGE_SIZE_RANKING)

    await callback.message.edit_text(
        text,
        reply_markup=keyboards.get_ranking_keyboard("teacher", page, total_pages, group_id)
    )
    await callback.answer()


@router.callback_query(F.data == "teacher:menu")
async def back_to_teacher_menu(callback: CallbackQuery):
    """Return to teacher menu"""
    await callback.message.delete()
    await safe_answer_callback(callback)


@router.callback_query(F.data == "teacher:pending")
async def show_pending_callback(callback: CallbackQuery):
    """Show pending approvals from callback"""
    pending = db.get_pending_approvals()

    if not pending:
        await safe_edit_message(
            callback,
            "✅ Hozircha tasdiq kutayotgan o'quvchi yo'q!",
            reply_markup=keyboards.get_back_keyboard("teacher:menu")
        )
        await callback.answer()
        return

    await safe_edit_message(
        callback,
        f"⏳ TASDIQ KUTAYOTGANLAR ({len(pending)})\n\n"
        f"🆕 — yangi ro'yxat\n"
        f"🔄 — tiklash so'rovi\n\n"
        f"O'quvchini tanlang:",
        reply_markup=keyboards.get_pending_list_keyboard(pending, page=0)
    )
    await callback.answer()


@router.message(F.text.contains("Pending"))
async def show_pending(message: Message):
    """Show pending approvals list"""
    pending = db.get_pending_approvals()

    if not pending:
        await message.answer(
            "✅ Hozircha tasdiq kutayotgan o'quvchi yo'q!",
            reply_markup=keyboards.get_teacher_keyboard(pending_count=0)
        )
        return

    await message.answer(
        f"⏳ TASDIQ KUTAYOTGANLAR ({len(pending)})\n\n"
        f"🆕 — yangi ro'yxat\n"
        f"🔄 — tiklash so'rovi\n\n"
        f"O'quvchini tanlang:",
        reply_markup=keyboards.get_pending_list_keyboard(pending, page=0)
    )


@router.callback_query(F.data.startswith("pending_detail:"))
async def show_pending_detail(callback: CallbackQuery):
    """Show pending student detail with approve/reject buttons"""
    user_id = callback.data.split(":")[1]
    user = db.get_user(user_id)

    if not user:
        await callback.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)
        return

    status = user.get('status', '')
    is_restore = status == 'pending_restore'

    group_id = user.get('group_id', '')
    group = db.get_group(group_id) if group_id else None
    group_name = group.get('name', group_id) if group else "Yo'q"

    icon = "🔄 TIKLASH SO'ROVI" if is_restore else "🆕 YANGI RO'YXAT"

    text = (
        f"{icon}\n\n"
        f"👤 Ism: {user.get('full_name', 'N/A')}\n"
        f"📱 Telefon: {user.get('phone', 'N/A')}\n"
        f"🔗 Username: @{user.get('username', 'N/A')}\n"
        f"👥 Guruh: {group_name}\n"
        f"🆔 ID: {user_id}\n"
    )

    if is_restore:
        text += f"\n⚠️ Oldingi ball: {user.get('points', 0)}"
        await callback.message.edit_text(
            text,
            reply_markup=keyboards.get_restore_approval_keyboard_with_back(user_id)
        )
    else:
        await callback.message.edit_text(
            text,
            reply_markup=keyboards.get_approval_keyboard(user_id, from_pending=True)
        )
    await callback.answer()


@router.callback_query(F.data.startswith("pending_page:"))
async def pending_page_handler(callback: CallbackQuery):
    """Handle pending list pagination"""
    page = int(callback.data.split(":")[1])
    pending = db.get_pending_approvals()

    if not pending:
        await safe_edit_message(
            callback,
            "✅ Hozircha tasdiq kutayotgan o'quvchi yo'q!",
            reply_markup=keyboards.get_back_keyboard("teacher:menu")
        )
        await callback.answer()
        return

    await safe_edit_message(
        callback,
        f"⏳ TASDIQ KUTAYOTGANLAR ({len(pending)})\n\n"
        f"🆕 — yangi ro'yxat\n"
        f"🔄 — tiklash so'rovi\n\n"
        f"O'quvchini tanlang:",
        reply_markup=keyboards.get_pending_list_keyboard(pending, page=page)
    )
    await callback.answer()


# ═══════════════════════════════════════════════════════════════════════════════
# SETTINGS HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data.startswith("settings:"))
async def handle_settings(callback: CallbackQuery):
    """Handle settings menu actions"""
    action = callback.data.split(":")[1]

    if action == "back":
        await callback.message.delete()
        await safe_answer_callback(callback)
        return

    if action == "groups":
        teacher_id = str(callback.from_user.id)
        groups = db.get_teacher_groups(teacher_id)

        text = "👥 GROUP MANAGEMENT\n\n"
        if groups:
            text += f"You have {len(groups)} group(s):\n"
            for group in groups:
                text += f"  • {group['name']} ({group['sheet_name']})\n"
        else:
            text += "You don't have any groups yet.\nCreate your first group to get started!"

        await safe_edit_message(
            callback,
            text,
            reply_markup=keyboards.get_groups_management_keyboard(teacher_id)
        )

    elif action == "commission":
        settings = db.get_settings()
        commission_rate = settings.get('commission_rate', config.DEFAULT_COMMISSION_RATE)
        commission_pool = int(settings.get('commission_pool', 0) or 0)

        await safe_edit_message(
            callback,
            f"TRANSFER COMMISSION\n"
            f"Current Rate: {commission_rate * 100}%\n"
            f"Commission Pool: {commission_pool} pts\n\n"
            f"Select new commission rate:",
            reply_markup=keyboards.get_commission_keyboard()
        )

    elif action == "transfer_limits":
        await safe_edit_message(
            callback,
            format_transfer_limits_text(),
            reply_markup=keyboards.get_transfer_limits_keyboard(db.get_settings())
        )

    elif action == "bot_status":
        settings = db.get_settings()
        current_status = settings.get('bot_status', 'public')

        status_description = {
            'public': '✅ Normal mode - All users can access bot',
            'maintenance': '🔧 Maintenance mode - Only teachers can access'
        }

        await safe_edit_message(
            callback,
            f"🔓 BOT STATUS\n"
            f"Current: {current_status.upper()}\n"
            f"{status_description.get(current_status, '')}\n\n"
            f"Select new status:",
            reply_markup=keyboards.get_bot_status_keyboard(current_status)
        )

    elif action == "sync_control":
        settings = db.get_settings()
        sync_enabled = settings.get('sync_enabled', True)
        sync_interval = int(settings.get('sync_interval', config.DEFAULT_SYNC_INTERVAL))
        status_text = 'ON' if sync_enabled else 'OFF'

        await safe_edit_message(
            callback,
            f"CACHE SETTINGS\n"
            f"Status: {status_text}\n"
            f"Interval: {sync_interval} seconds\n\n"
            f"Auto Sync ON: cache refreshes in background.\n"
            f"Auto Sync OFF: first read loads from Sheets, next reads use cached data until manual refresh or changes.",
            reply_markup=keyboards.get_sync_control_keyboard(sync_enabled)
        )

    elif action == "transaction_history":
        await safe_edit_message(
            callback,
            "🧾 TRANSACTION HISTORY\nFilter transaction logs:",
            reply_markup=keyboards.get_transaction_history_keyboard()
        )

    elif action == "export":
        await safe_edit_message(
            callback,
            f"📥 EXPORT DATA\n"
            f"Select export format:",
            reply_markup=keyboards.get_export_keyboard()
        )

    elif action == "edit_rules":
        # Get current rules from database
        settings = db.get_settings()
        current_rules = settings.get('rules_text', 'No rules set yet.')

        await safe_edit_message(
            callback,
            f"📝 BOT RULES\n\n"
            f"Current Rules:\n"
            f"{current_rules}\n\n"
            f"Click below to edit rules:",
            reply_markup=keyboards.get_edit_rules_keyboard()
        )

    elif action == "broadcast":
        await safe_edit_message(
            callback,
            f"📢 GLOBAL BROADCAST\n"
            f"Select target audience:",
            reply_markup=keyboards.get_broadcast_keyboard()
        )

    await safe_answer_callback(callback)


@router.callback_query(F.data.startswith("sync:"))
async def handle_sync_settings(callback: CallbackQuery):
    """Handle Sheets cache settings."""
    parts = callback.data.split(":")
    action = parts[1]
    settings = db.get_settings()
    sync_enabled = settings.get('sync_enabled', True)
    sync_interval = int(settings.get('sync_interval', config.DEFAULT_SYNC_INTERVAL))

    if action == "control":
        await safe_edit_message(
            callback,
            f"CACHE SETTINGS\n"
            f"Status: {'ON' if sync_enabled else 'OFF'}\n"
            f"Interval: {sync_interval} seconds\n\n"
            f"Auto Sync ON: cache refreshes in background.\n"
            f"Auto Sync OFF: first read loads from Sheets, next reads use cached data until manual refresh or changes.",
            reply_markup=keyboards.get_sync_control_keyboard(sync_enabled)
        )
        await safe_answer_callback(callback)
        return

    if action == "toggle":
        new_enabled = not sync_enabled
        if not db.update_settings({'sync_enabled': new_enabled}):
            await safe_answer_callback(callback, "Update failed", show_alert=True)
            return
        sheets_manager.configure_cache_policy(new_enabled, sync_interval)
        if new_enabled:
            sheets_manager.start_background_sync()
        else:
            sheets_manager.stop_background_sync()
        await safe_answer_callback(callback, f"Cache {'enabled' if new_enabled else 'disabled'}")
        await safe_edit_message(
            callback,
            f"CACHE SETTINGS\n"
            f"Status: {'ON' if new_enabled else 'OFF'}\n"
            f"Interval: {sync_interval} seconds\n\n"
            f"Auto Sync ON: cache refreshes in background.\n"
            f"Auto Sync OFF: first read loads from Sheets, next reads use cached data until manual refresh or changes.",
            reply_markup=keyboards.get_sync_control_keyboard(new_enabled)
        )
        return

    if action == "interval":
        await safe_edit_message(
            callback,
            f"CACHE INTERVAL\n"
            f"Current interval: {sync_interval} seconds\n\n"
            f"Select new interval:",
            reply_markup=keyboards.get_sync_interval_keyboard()
        )
        await safe_answer_callback(callback)
        return

    if action == "set_interval":
        interval = int(parts[2])
        if not db.update_settings({'sync_interval': interval}):
            await safe_answer_callback(callback, "Update failed", show_alert=True)
            return
        sheets_manager.configure_cache_policy(sync_enabled, interval)
        if sync_enabled and not sheets_manager.is_sync_running():
            sheets_manager.start_background_sync()
        await safe_answer_callback(callback, f"Interval set to {interval}s")
        await safe_edit_message(
            callback,
            f"CACHE SETTINGS\n"
            f"Status: {'ON' if sync_enabled else 'OFF'}\n"
            f"Interval: {interval} seconds\n\n"
            f"Auto Sync ON: cache refreshes in background.\n"
            f"Auto Sync OFF: first read loads from Sheets, next reads use cached data until manual refresh or changes.",
            reply_markup=keyboards.get_sync_control_keyboard(sync_enabled)
        )
        return

    if action == "refresh":
        sheets_manager.invalidate_cache()
        await safe_answer_callback(callback, "Cache cleared")
        await safe_edit_message(
            callback,
            f"CACHE SETTINGS\n"
            f"Status: {'ON' if sync_enabled else 'OFF'}\n"
            f"Interval: {sync_interval} seconds\n\n"
            f"Cache cleared. Next read will fetch fresh data from Sheets.",
            reply_markup=keyboards.get_sync_control_keyboard(sync_enabled)
        )
        return

    await safe_answer_callback(callback)


@router.callback_query(F.data.startswith("bot_status:"))
async def change_bot_status(callback: CallbackQuery):
    """Change bot status"""
    new_status = callback.data.split(":")[1]

    if not db.update_settings({'bot_status': new_status}):
        await safe_edit_message(
            callback,
            "Failed to update bot status. Please try again.",
            reply_markup=keyboards.get_back_keyboard("settings:back")
        )
        await safe_answer_callback(callback, "Update failed", show_alert=True)
        return

    status_info = {
        'public': '✅ PUBLIC MODE\nAll users can access the bot normally.',
        'maintenance': '🔧 MAINTENANCE MODE\nOnly teachers can access. Students will see maintenance message.'
    }

    await safe_edit_message(
        callback,
        f"✅ Bot status updated!\n\n"
        f"{status_info.get(new_status, '')}",
        reply_markup=keyboards.get_back_keyboard("settings:back")
    )
    await safe_answer_callback(callback, f"✅ Changed to {new_status.upper()} mode")



@router.callback_query(F.data.startswith("logs:"))
async def handle_transaction_logs(callback: CallbackQuery):
    """Handle transaction logs"""
    parts = callback.data.split(":")
    action = parts[1]

    if action == "export_menu":
        # Show export format menu
        await safe_edit_message(
            callback,
            f"📊 EXPORT TRANSACTION LOGS\n"
            f"Select export format:",
            reply_markup=keyboards.get_logs_export_keyboard()
        )
        await safe_answer_callback(callback)
        return

    if action == "export":
        # Handle export based on format
        export_format = parts[2] if len(parts) > 2 else "excel"
        print(f"📊 Export requested: {export_format}")
        await callback.answer(f"📥 Generating {export_format.upper()} export...")

        try:
            # Get all logs
            logs = db.get_transaction_logs(limit=500)  # Get more logs for export
            print(f"📋 Found {len(logs)} logs to export")

            if not logs:
                await callback.message.edit_text(
                    "📜 No transaction logs found to export.",
                    reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
                )
                await callback.answer()
                return

            from aiogram.types import FSInputFile
            import os

            if export_format == "excel":
                # Generate Excel export
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                filename = f"transaction_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                wb = Workbook()
                ws = wb.active
                ws.title = "Transaction Logs"

                # Header style
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                # Headers
                headers = ['Date', 'Type', 'Description', 'Amount', 'Commission', 'Status']
                ws.append(headers)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Data
                for log in logs:
                    log_type = log.get('type', 'unknown')
                    timestamp = log.get('timestamp', 'N/A')

                    if log_type == "transfer":
                        description = (
                            f"From: {log.get('sender_name', 'Unknown')} -> To: {log.get('recipient_name', 'Unknown')} | "
                            f"Sender: {log.get('sender_old_balance', 'N/A')}->{log.get('sender_new_balance', 'N/A')} | "
                            f"Recipient: {log.get('recipient_old_balance', 'N/A')}->{log.get('recipient_new_balance', 'N/A')}"
                        )
                        amount = log.get('amount', 0)
                        commission = log.get('commission', 0)
                    elif log_type == "add_points":
                        description = (
                            f"Added to: {log.get('student_name', 'Unknown')} | "
                            f"Balance: {log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}"
                        )
                        amount = log.get('amount', 0)
                        commission = 0
                    elif log_type == "subtract_points":
                        description = (
                            f"Subtracted from: {log.get('student_name', 'Unknown')} | "
                            f"Balance: {log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}"
                        )
                        amount = -log.get('amount', 0)
                        commission = 0
                    else:
                        description = "Unknown transaction"
                        amount = 0
                        commission = 0

                    ws.append([
                        str(timestamp),
                        log_type.replace('_', ' ').title(),
                        description,
                        amount,
                        commission,
                        log.get('status', 'completed')
                    ])

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(filename)

                await callback.message.answer_document(
                    FSInputFile(filename),
                    caption=(
                        f"📋 Transaction Logs Export (Excel)\n"
                        f"Total: {len(logs)} transactions\n"
                        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                )
                os.remove(filename)

            elif export_format == "pdf":
                # Generate PDF export
                from reportlab.lib.pagesizes import A4, landscape
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.enums import TA_CENTER

                filename = f"transaction_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

                doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
                elements = []
                styles = getSampleStyleSheet()

                # Title
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=24,
                    textColor=colors.HexColor('#366092'),
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                title = Paragraph("Transaction Logs Report", title_style)
                elements.append(title)

                # Summary
                summary_text = (
                    f"<b>Export Date:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
                    f"<b>Total Transactions:</b> {len(logs)}<br/>"
                    f"<b>Report Period:</b> All time"
                )
                summary = Paragraph(summary_text, styles['Normal'])
                elements.append(summary)
                elements.append(Spacer(1, 0.3*inch))

                # Table data
                data = [['Date', 'Type', 'Description', 'Amount', 'Status']]

                for log in logs[:100]:  # Limit to 100 for PDF
                    log_type = log.get('type', 'unknown')
                    timestamp = str(log.get('timestamp', 'N/A'))[:19]

                    if log_type == "transfer":
                        description = (
                            f"{log.get('sender_name', 'Unknown')[:12]}->{log.get('recipient_name', 'Unknown')[:12]} | "
                            f"S:{log.get('sender_old_balance', 'N/A')}->{log.get('sender_new_balance', 'N/A')} | "
                            f"R:{log.get('recipient_old_balance', 'N/A')}->{log.get('recipient_new_balance', 'N/A')}"
                        )
                        amount = f"+{log.get('amount', 0)}"
                    elif log_type == "add_points":
                        description = (
                            f"Added: {log.get('student_name', 'Unknown')[:12]} | "
                            f"{log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}"
                        )
                        amount = f"+{log.get('amount', 0)}"
                    elif log_type == "subtract_points":
                        description = (
                            f"Removed: {log.get('student_name', 'Unknown')[:12]} | "
                            f"{log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}"
                        )
                        amount = f"-{log.get('amount', 0)}"
                    else:
                        description = "Unknown"
                        amount = "0"

                    data.append([
                        timestamp,
                        log_type.replace('_', ' ').title(),
                        description,
                        amount,
                        log.get('status', 'completed')[:10]
                    ])

                # Create table
                table = Table(data, colWidths=[1.8*inch, 1.2*inch, 3*inch, 0.8*inch, 1*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
                ]))

                elements.append(table)

                if len(logs) > 100:
                    elements.append(Spacer(1, 0.2*inch))
                    note = Paragraph(f"<i>Note: Showing first 100 of {len(logs)} transactions</i>", styles['Normal'])
                    elements.append(note)

                doc.build(elements)

                await callback.message.answer_document(
                    FSInputFile(filename),
                    caption=(
                        f"📈 Transaction Logs Export (PDF)\n"
                        f"Total: {len(logs)} transactions\n"
                        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    )
                )
                os.remove(filename)

            await callback.message.edit_text(
                f"✅ Transaction logs exported successfully as {export_format.upper()}!\n"
                f"Total: {len(logs)} transactions",
                reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
            )

        except Exception as e:
            await callback.message.edit_text(
                f"❌ Error exporting logs:\n{str(e)}",
                reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
            )

        return

    # Get logs based on filter
    await callback.answer("🔍 Loading logs...")

    try:
        if action == "all":
            logs = db.get_transaction_logs(limit=config.TRANSACTION_LOG_LIMIT)
            filter_name = "ALL"
        elif action == "transfer":
            logs = db.get_transaction_logs(limit=config.TRANSACTION_LOG_LIMIT, transaction_type="transfer")
            filter_name = "TRANSFERS"
        elif action == "add_points":
            logs = db.get_transaction_logs(limit=config.TRANSACTION_LOG_LIMIT, transaction_type="add_points")
            filter_name = "ADDED POINTS"
        elif action == "subtract_points":
            logs = db.get_transaction_logs(limit=config.TRANSACTION_LOG_LIMIT, transaction_type="subtract_points")
            filter_name = "SUBTRACTED POINTS"
        elif action == "clear":
            # Clear all transaction logs
            await safe_edit_message(
                callback,
                "⚠️ CLEAR ALL TRANSACTION LOGS\n\n"
                "Are you sure you want to delete ALL transaction logs?\n"
                "This action cannot be undone!",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="✅ Yes, Clear All", callback_data="logs:clear_confirm")],
                    [InlineKeyboardButton(text="❌ Cancel", callback_data="settings:transaction_history")]
                ])
            )
            await safe_answer_callback(callback)
            return
        elif action == "clear_confirm":
            # Confirmed - delete all logs with progress
            await safe_edit_message(
                callback,
                "🗑️ CLEARING LOGS...\n\n"
                "⏳ Starting deletion process...\n"
                "Please wait...",
                reply_markup=None
            )
            await safe_answer_callback(callback)

            # Progress callback function
            async def update_progress(deleted, total, progress):
                try:
                    bar_length = 20
                    filled = int(bar_length * progress / 100)
                    bar = "█" * filled + "░" * (bar_length - filled)

                    await callback.message.edit_text(
                        f"🗑️ CLEARING LOGS...\n\n"
                        f"Progress: {progress}%\n"
                        f"[{bar}]\n\n"
                        f"Deleted: {deleted}/{total} logs"
                    )
                except:
                    pass  # Ignore telegram rate limit errors

            # Delete with progress
            deleted_count = db.clear_all_transaction_logs(progress_callback=update_progress)

            # Final message
            await safe_edit_message(
                callback,
                f"✅ LOGS CLEARED\n\n"
                f"Deleted {deleted_count} transaction log(s) from Firestore.",
                reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
            )
            return
        else:
            await callback.answer("❌ Unknown action", show_alert=True)
            return

        if not logs:
            await callback.message.edit_text(
                f"📜 No transaction logs found for: {filter_name}",
                reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
            )
            return

        # Format logs
        text = f"📜 TRANSACTION LOGS ({filter_name})\n"

        for log in logs[:15]:
            log_type = log.get('type', 'unknown')

            if log_type == "transfer":
                text += (
                    f"💸 Transfer\n"
                    f"From: {log.get('sender_name', 'Unknown')}\n"
                    f"To: {log.get('recipient_name', 'Unknown')}\n"
                    f"Amount: {log.get('amount', 0)} pts\n"
                    f"Commission: {log.get('commission', 0)} pts\n\n"
                )
            elif log_type == "add_points":
                text += (
                    f"➕ Added Points\n"
                    f"Student: {log.get('student_name', 'Unknown')}\n"
                    f"Amount: +{log.get('amount', 0)} pts\n"
                    f"Balance: {log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}\n\n"
                )
            elif log_type == "subtract_points":
                text += (
                    f"➖ Subtracted Points\n"
                    f"Student: {log.get('student_name', 'Unknown')}\n"
                    f"Amount: -{log.get('amount', 0)} pts\n"
                    f"Balance: {log.get('old_balance', 'N/A')}->{log.get('new_balance', 'N/A')}\n\n"
                )

        if len(logs) > 15:
            text += f"\n... and {len(logs) - 15} more transactions\n"

        text += f"\nTotal: {len(logs)} transactions"

        await callback.message.edit_text(
            text,
            reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
        )

    except Exception as e:
        await callback.message.edit_text(
            f"❌ Error loading transaction logs:\n{str(e)}",
            reply_markup=keyboards.get_back_keyboard("settings:transaction_history")
        )


@router.callback_query(F.data.startswith("commission:"))
async def handle_commission(callback: CallbackQuery):
    """Handle commission rate changes"""
    action = callback.data.split(":")[1]

    if action == "set":
        rate = int(callback.data.split(":")[2])
        rate_decimal = rate / 100.0

        # Validate rate
        if rate < 0 or rate > 50:
            await callback.answer("❌ Rate must be between 0-50%")
            return

        # Update settings
        if not db.update_settings({'commission_rate': rate_decimal}):
            await callback.answer("Failed to update commission rate", show_alert=True)
            return

        await callback.answer(f"✅ Commission rate set to {rate}%")

        await callback.message.edit_text(
            f"✅ COMMISSION UPDATED\n"
            f"New Rate: {rate}%\n\n"
            f"This rate will apply to all future transfers.",
            reply_markup=keyboards.get_back_keyboard("settings:back")
        )


@router.callback_query(F.data.startswith("transfer_limits:"))
async def handle_transfer_limits(callback: CallbackQuery, state: FSMContext):
    """Handle transfer limit settings."""
    parts = callback.data.split(":")
    action = parts[1]

    labels = {
        'daily_transfer_count_limit': 'daily transfer count limit',
        'weekly_transfer_count_limit': 'weekly transfer count limit',
        'daily_transfer_points_limit': 'daily transfer points limit',
        'weekly_transfer_points_limit': 'weekly transfer points limit',
    }

    if action == "set":
        setting_key = parts[2]
        if setting_key not in labels:
            await safe_answer_callback(callback, "Unknown limit", show_alert=True)
            return

        await state.set_state(SettingsStates.waiting_for_transfer_limit)
        await state.update_data(
            transfer_limit_key=setting_key,
            transfer_limit_label=labels[setting_key]
        )
        await callback.message.edit_text(
            f"Enter new value for {labels[setting_key]}.\n\n"
            f"Send 0 for unlimited.\n"
            f"Only whole numbers are allowed.\n\n"
            f"Send /cancel to abort."
        )
        await safe_answer_callback(callback)
        return

    if action == "reset_usage":
        if not db.reset_all_transfer_usage():
            await safe_answer_callback(callback, "Reset failed", show_alert=True)
            return

        await safe_edit_message(
            callback,
            format_transfer_limits_text() + "\n\nAll tracked transfer usage counters were reset.",
            reply_markup=keyboards.get_transfer_limits_keyboard(db.get_settings())
        )
        await safe_answer_callback(callback, "Transfer usage reset")
        return

    await safe_answer_callback(callback)


@router.message(SettingsStates.waiting_for_transfer_limit)
async def process_transfer_limit_value(message: Message, state: FSMContext):
    """Save numeric transfer limit setting."""
    raw_value = (message.text or "").strip()
    try:
        value = int(raw_value)
        if value < 0:
            raise ValueError
    except ValueError:
        await message.answer("Please send a whole number 0 or greater.")
        return

    data = await state.get_data()
    setting_key = data.get('transfer_limit_key')
    label = data.get('transfer_limit_label', 'transfer limit')
    target_user_id = data.get('transfer_limit_target_user_id')
    target_name = data.get('transfer_limit_target_name')

    if not setting_key:
        await message.answer("Transfer limit setup state was lost. Open settings again.")
        await state.clear()
        return

    if target_user_id:
        if not db.update_transfer_limit_override(target_user_id, {setting_key: value}):
            await message.answer("Failed to save the per-user transfer limit.")
            return
    else:
        if not db.update_settings({setting_key: value}):
            await message.answer("Failed to save the transfer limit.")
            return

    await state.clear()
    if target_user_id and target_name:
        await message.answer(
            f"Updated {target_name} {label} to {'global' if value == 0 else value}.",
            reply_markup=keyboards.get_back_keyboard(f"student_transfer_limits:view:{target_user_id}")
        )
        return

    await message.answer(
        f"Updated {label} to {'unlimited' if value == 0 else value}.",
        reply_markup=keyboards.get_back_keyboard("settings:transfer_limits")
    )


@router.callback_query(F.data.startswith("broadcast:"))
async def handle_broadcast(callback: CallbackQuery, state: FSMContext):
    """Handle broadcast message"""
    target = callback.data.split(":")[1]

    await state.set_state(BroadcastStates.waiting_for_message)
    await state.update_data(target=target)

    target_text = {
        'all_active': '👥 All Active Users',
        'students': '👨‍🎓 Students Only',
        'teachers': '👨‍🏫 Teachers Only'
    }.get(target, 'Unknown')

    await callback.message.edit_text(
        f"📢 BROADCAST MESSAGE\n"
        f"Target: {target_text}\n\n"
        f"Send your message now:\n"
        f"(Text, photo, video, or document)\n\n"
        f"Send /cancel to abort."
    )
    await callback.answer()


@router.message(BroadcastStates.waiting_for_message)
async def process_broadcast_message(message: Message, state: FSMContext):
    """Process and send broadcast message"""
    data = await state.get_data()
    target = data.get('target')

    # Get target users
    if target == 'all_active':
        users = db.get_all_users(status='active')
    elif target == 'students':
        users = db.get_all_users(role='student', status='active')
    elif target == 'teachers':
        users = db.get_all_users(role='teacher', status='active')
    else:
        await message.answer("❌ Invalid target")
        await state.clear()
        return

    # Send broadcast
    success_count = 0
    fail_count = 0

    status_msg = await message.answer(f"📢 Broadcasting to {len(users)} users...")

    for user in users:
        try:
            if message.text:
                await message.bot.send_message(user['user_id'], f"📢 Broadcast:\n\n{message.text}")
            elif message.photo:
                await message.bot.send_photo(user['user_id'], message.photo[-1].file_id, caption=message.caption)
            elif message.video:
                await message.bot.send_video(user['user_id'], message.video.file_id, caption=message.caption)
            elif message.document:
                await message.bot.send_document(user['user_id'], message.document.file_id, caption=message.caption)

            success_count += 1
        except Exception as e:
            fail_count += 1
            print(f"Failed to send to {user['user_id']}: {e}")

    await status_msg.edit_text(
        f"✅ BROADCAST COMPLETE\n"
        f"✅ Sent: {success_count}\n"
        f"❌ Failed: {fail_count}\n"
        f"📊 Total: {len(users)}"
    )

    await state.clear()


@router.callback_query(F.data.startswith("rules:"))
async def handle_rules(callback: CallbackQuery, state: FSMContext):
    """Handle rules editing"""
    action = callback.data.split(":")[1]

    if action == "edit":
        # Start edit rules flow
        settings = db.get_settings()
        current_rules = settings.get('rules_text', 'No rules set yet.')

        await state.set_state(EditRulesStates.waiting_for_rules)

        await callback.message.edit_text(
            f"📝 EDIT BOT RULES\n\n"
            f"Current Rules:\n"
            f"{current_rules}\n\n"
            f"Send new rules text now.\n"
            f"You can use formatting:\n"
            f"• Line breaks for new lines\n"
            f"• Emojis 😊\n"
            f"• Keep it clear and concise\n\n"
            f"Send /cancel to abort."
        )
        await callback.answer()


@router.message(EditRulesStates.waiting_for_rules)
async def process_rules_text(message: Message, state: FSMContext):
    """Process new rules text"""
    new_rules = message.text.strip()

    if len(new_rules) < 10:
        await message.answer("❌ Rules text is too short (minimum 10 characters). Try again:")
        return

    if len(new_rules) > 2000:
        await message.answer("❌ Rules text is too long (maximum 2000 characters). Try again:")
        return

    # Save to database
    if not db.update_settings({'rules_text': new_rules}):
        await message.answer("Failed to update rules. Please try again.")
        return

    await message.answer(
        f"✅ RULES UPDATED SUCCESSFULLY!\n\n"
        f"New Rules:\n"
        f"{new_rules[:500]}{'...' if len(new_rules) > 500 else ''}\n\n"
        f"Students will see these rules when they click 'Rules' button."
    )

    await state.clear()


@router.callback_query(F.data.startswith("compare:"))
async def handle_compare(callback: CallbackQuery):
    """Handle data comparison"""
    action = callback.data.split(":")[1]

    if action == "refresh":
        await callback.answer("🔍 Comparing data...")

        # Get Sheets data from current source of truth
        fb_users = db.get_all_users(role='student')

        # Get Sheets data
        sheet_data = await sheets_manager.get_all_users_from_sheets()

        # Compare
        fb_ids = {u['user_id'] for u in fb_users}
        sheet_ids = {u['user_id'] for u in sheet_data}

        only_fb = fb_ids - sheet_ids
        only_sheet = sheet_ids - fb_ids
        common = fb_ids & sheet_ids

        # Check differences in points
        differences = []
        for user_id in common:
            fb_user = next(u for u in fb_users if u['user_id'] == user_id)
            sheet_user = next(u for u in sheet_data if u['user_id'] == user_id)

            fb_points = fb_user.get('points', 0)
            sheet_points = sheet_user.get('points', 0)

            if fb_points != sheet_points:
                differences.append({
                    'user_id': user_id,
                    'name': fb_user.get('full_name', 'Unknown'),
                    'fb_points': fb_points,
                    'sheet_points': sheet_points,
                    'diff': fb_points - sheet_points
                })

        text = f"🔍 DATA COMPARISON\n"
        text += f"📊 Statistics:\n"
        text += f"• Common: {len(common)}\n"
        text += f"• Only in Sheets cache: {len(only_fb)}\n"
        text += f"• Only in Sheets: {len(only_sheet)}\n"
        text += f"• Point differences: {len(differences)}\n\n"

        if differences:
            text += f"⚠️ Points Mismatch:\n"
            for diff in differences[:5]:
                text += f"• {diff['name']}: FB={diff['fb_points']}, Sheet={diff['sheet_points']}\n"

            if len(differences) > 5:
                text += f"\n... and {len(differences) - 5} more\n"

        await callback.message.edit_text(
            text,
            reply_markup=keyboards.get_comparison_keyboard()
        )

    elif action in ("sync_fb_to_sh", "sync_sh_to_fb"):
        await callback.answer("Sync is disabled in Sheets-only mode.", show_alert=True)
        await callback.message.edit_text(
            "🔄 Sync is disabled in Sheets-only mode.\n\nUse the Sheet tabs as the source of truth.",
            reply_markup=keyboards.get_comparison_keyboard()
        )

    elif action == "export":
        await callback.answer("📥 Generating comparison report...")

        try:
            # Get Sheets data from current source of truth
            fb_users = db.get_all_users(role='student')

            # Get Sheets data
            sheet_data = await sheets_manager.get_all_users_from_sheets()

            # Compare
            fb_ids = {u['user_id'] for u in fb_users}
            sheet_ids = {u['user_id'] for u in sheet_data}

            only_fb = fb_ids - sheet_ids
            only_sheet = sheet_ids - fb_ids
            common = fb_ids & sheet_ids

            # Check differences in points
            differences = []
            for user_id in common:
                fb_user = next(u for u in fb_users if u['user_id'] == user_id)
                sheet_user = next(u for u in sheet_data if u['user_id'] == user_id)

                fb_points = fb_user.get('points', 0)
                sheet_points = sheet_user.get('points', 0)

                if fb_points != sheet_points:
                    differences.append({
                        'user_id': user_id,
                        'name': fb_user.get('full_name', 'Unknown'),
                        'fb_points': fb_points,
                        'sheet_points': sheet_points,
                        'diff': fb_points - sheet_points
                    })

            # Generate detailed report as JSON
            import json
            report_data = {
                "report_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "statistics": {
                    "total_common": len(common),
                    "only_in_source": len(only_fb),
                    "only_in_sheets": len(only_sheet),
                    "point_differences": len(differences)
                },
                "users_only_in_source": [
                    {
                        "user_id": uid,
                        "name": next((u['full_name'] for u in fb_users if u['user_id'] == uid), "Unknown"),
                        "points": next((u['points'] for u in fb_users if u['user_id'] == uid), 0)
                    } for uid in only_fb
                ],
                "users_only_in_sheets": [
                    {
                        "user_id": uid,
                        "name": next((u['full_name'] for u in sheet_data if u['user_id'] == uid), "Unknown"),
                        "points": next((u['points'] for u in sheet_data if u['user_id'] == uid), 0)
                    } for uid in only_sheet
                ],
                "point_mismatches": differences
            }

            # Save to file
            filename = f"comparison_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)

            # Send file
            from aiogram.types import FSInputFile
            await callback.message.answer_document(
                FSInputFile(filename),
                caption=(
                    f"📊 Data Comparison Report\n"
                    f"• Common users: {len(common)}\n"
                    f"• Only in Sheets cache: {len(only_fb)}\n"
                    f"• Only in Sheets: {len(only_sheet)}\n"
                    f"• Point differences: {len(differences)}"
                )
            )

            # Delete temp file
            import os
            os.remove(filename)

            await callback.message.edit_text(
                f"✅ Comparison report exported successfully!",
                reply_markup=keyboards.get_comparison_keyboard()
            )

        except Exception as e:
            await callback.message.edit_text(
                f"❌ Error exporting comparison report:\n{str(e)}",
                reply_markup=keyboards.get_comparison_keyboard()
            )


@router.callback_query(F.data.startswith("export:"))
async def handle_export(callback: CallbackQuery):
    """Handle data export"""
    format_type = callback.data.split(":")[1]

    if format_type == "sheets_copy":
        await callback.answer("🔗 Opening Google Sheets...")
        sheet_url = f"https://docs.google.com/spreadsheets/d/{config.SHEET_ID}"
        await callback.message.edit_text(
            f"📊 GOOGLE SHEETS\n"
            f"Access your data:\n\n"
            f"🔗 {sheet_url}\n\n"
            f"Use File → Make a copy to create your own version.",
            reply_markup=keyboards.get_back_keyboard("settings:export")
        )
        return

    await callback.answer(f"📥 Preparing {format_type.upper()} export...")

    # Get all users
    users = db.get_all_users(role='student', status='active')

    if format_type == "json":
        import json
        data = {
            "export_date": datetime.now().isoformat(),
            "total_students": len(users),
            "students": users
        }

        # Save to file
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

        # Send file
        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            FSInputFile(filename),
            caption=f"📥 JSON Export\nTotal: {len(users)} students"
        )

        # Delete temp file
        import os
        os.remove(filename)

        await callback.message.edit_text(
            f"✅ JSON export complete!",
            reply_markup=keyboards.get_back_keyboard("settings:export")
        )


    elif format_type == "excel":
        from openpyxl import Workbook
        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Headers
        ws.append(['User ID', 'Full Name', 'Username', 'Phone', 'Points', 'Status'])

        # Data
        for user in users:
            ws.append([
                user.get('user_id', ''),
                user.get('full_name', ''),
                user.get('username', ''),
                user.get('phone', ''),
                user.get('points', 0),
                user.get('status', '')
            ])

        wb.save(filename)

        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            FSInputFile(filename),
            caption=f"📥 Excel Export\nTotal: {len(users)} students"
        )

        import os
        os.remove(filename)

        await callback.message.edit_text(
            f"✅ Excel export complete!",
            reply_markup=keyboards.get_back_keyboard("settings:export")
        )

    elif format_type == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        filename = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("Points Management System - Student Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))

        # Summary
        summary_text = f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Students: {len(users)}"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 0.3*inch))

        # Table data
        data = [['#', 'Name', 'Username', 'Points', 'Status']]
        for idx, user in enumerate(users, 1):
            data.append([
                str(idx),
                user.get('full_name', '')[:20],
                user.get('username', '')[:15],
                str(user.get('points', 0)),
                user.get('status', '')
            ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)

        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            FSInputFile(filename),
            caption=f"📥 PDF Report\nTotal: {len(users)} students"
        )

        import os
        os.remove(filename)

        await callback.message.edit_text(
            f"✅ PDF export complete!",
            reply_markup=keyboards.get_back_keyboard("settings:export")
        )


# ═══════════════════════════════════════════════════════════════════════════════
# GROUP MANAGEMENT HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.callback_query(F.data.startswith("groups:"))
async def handle_groups_actions(callback: CallbackQuery, state: FSMContext):
    """Handle group management actions"""
    action = callback.data.split(":")[1]

    if action == "list":
        teacher_id = str(callback.from_user.id)
        groups = db.get_teacher_groups(teacher_id)

        if not groups:
            await safe_edit_message(
                callback,
                "You don't have any groups yet.\nCreate your first group!",
                reply_markup=keyboards.get_back_keyboard("settings:groups")
            )
            return

        text = "📋 YOUR GROUPS\n\n"
        for idx, group in enumerate(groups, 1):
            text += f"{idx}. {group['name']}\n"
            text += f"   📄 Sheet: {group['sheet_name']}\n"

            # Count students in this group
            students = db.get_all_users(role='student', status='active', group_id=group['group_id'])
            text += f"   👥 Students: {len(students)}\n\n"

        await safe_edit_message(
            callback,
            text,
            reply_markup=keyboards.get_groups_list_keyboard(groups, action="view")
        )

    elif action == "create":
        await safe_edit_message(
            callback,
            "📝 CREATE NEW GROUP\n\n"
            "Enter the group name (e.g., 'Class 10-A', 'Group B'):"
        )
        await state.set_state(GroupStates.waiting_for_name)
        await safe_answer_callback(callback)

    elif action == "refresh":
        # Refresh groups from Google Sheets
        teacher_id = str(callback.from_user.id)
        groups = db.get_teacher_groups(teacher_id, force_refresh=True)

        text = "👥 GROUP MANAGEMENT\n\n"
        if groups:
            text += f"✅ Refreshed! You have {len(groups)} group(s):\n"
            for group in groups:
                student_count = group.get('student_count', 0)
                text += f"  • {group['name']} ({student_count} students)\n"
        else:
            text += "No groups found.\nCreate tabs manually in Google Sheets!"

        await safe_edit_message(
            callback,
            text,
            reply_markup=keyboards.get_groups_management_keyboard(teacher_id)
        )
        await safe_answer_callback(callback, "✅ Groups refreshed!")

    elif action == "switch":
        teacher_id = str(callback.from_user.id)
        groups = db.get_teacher_groups(teacher_id)

        if not groups:
            await safe_edit_message(
                callback,
                "You don't have any groups to switch to.\nCreate a group first!",
                reply_markup=keyboards.get_back_keyboard("settings:groups")
            )
            return

        await safe_edit_message(
            callback,
            "🔄 SWITCH ACTIVE GROUP\n\n"
            "Select the group you want to work with:",
            reply_markup=keyboards.get_groups_list_keyboard(groups, action="switch")
        )


@router.message(GroupStates.waiting_for_name)
async def process_group_name(message: Message, state: FSMContext):
    """Process sheet name (which is also the group name)"""
    sheet_name = message.text.strip()

    if len(sheet_name) < 2:
        await message.answer("Sheet name is too short. Please enter at least 2 characters:")
        return

    teacher_id = str(message.from_user.id)

    # Check if sheet name already exists
    existing_sheets = sheets_manager.get_sheet_names()
    if sheet_name in existing_sheets:
        await message.answer(
            f"⚠️ Sheet '{sheet_name}' already exists!\n"
            f"Please enter a different name:"
        )
        return

    # Create the group (name = sheet_name)
    group_data = {
        'name': sheet_name,  # Same as sheet name
        'sheet_name': sheet_name,
        'teacher_id': teacher_id
    }

    group_id = db.create_group(group_data)

    if not group_id:
        await message.answer("❌ Failed to create group. Please try again.")
        await state.clear()
        return

    # Create the sheet tab
    success = sheets_manager.create_sheet_tab(sheet_name)

    if not success:
        # Rollback
        db.delete_group(group_id)
        await message.answer("❌ Failed to create Google Sheets tab. Please try again.")
        await state.clear()
        return

    # 🔄 Auto-refresh groups cache after creation
    db.get_teacher_groups(teacher_id, force_refresh=True)
    print(f"🔄 Auto-refreshed groups cache after group creation")

    await message.answer(
        f"✅ GROUP CREATED!\n\n"
        f"📄 Sheet: {sheet_name}\n\n"
        f"Students can now select this group during registration!",
        reply_markup=keyboards.get_teacher_keyboard()
    )

    await state.clear()


# Removed - sheet name is now entered directly in waiting_for_name handler


@router.callback_query(F.data.startswith("group_view:"))
async def handle_group_view(callback: CallbackQuery):
    """View group details"""
    group_id = callback.data.split(":")[1]
    group = db.get_group(group_id)

    if not group:
        await safe_answer_callback(callback, "Group not found!", show_alert=True)
        return

    # Count students
    students = db.get_all_users(role='student', status='active', group_id=group_id)

    text = f"📚 GROUP DETAILS\n\n"
    text += f"Name: {group['name']}\n"
    text += f"Sheet: {group['sheet_name']}\n"
    text += f"Students: {len(students)}\n"
    text += f"Status: {group.get('status', 'active')}\n"

    await safe_edit_message(
        callback,
        text,
        reply_markup=keyboards.get_group_detail_keyboard(group_id)
    )


@router.callback_query(F.data.startswith("group_students:"))
async def handle_group_students(callback: CallbackQuery):
    """View students in a group"""
    group_id = callback.data.split(":")[1]
    group = db.get_group(group_id)

    if not group:
        await safe_answer_callback(callback, "Group not found!", show_alert=True)
        return

    students = db.get_all_users(role='student', status='active', group_id=group_id)

    if not students:
        await safe_edit_message(
            callback,
            f"📚 {group['name']}\n\n"
            f"No students in this group yet.",
            reply_markup=keyboards.get_back_keyboard(f"group_view:{group_id}")
        )
        return

    # Sort by points
    students.sort(key=lambda x: x.get('points', 0), reverse=True)

    text = f"👥 STUDENTS IN {group['name']}\n\n"
    for idx, student in enumerate(students[:20], 1):  # Show top 20
        text += f"{idx}. {student['full_name']} - {student.get('points', 0)} pts\n"

    if len(students) > 20:
        text += f"\n... and {len(students) - 20} more"

    await safe_edit_message(
        callback,
        text,
        reply_markup=keyboards.get_back_keyboard(f"group_view:{group_id}")
    )


@router.callback_query(F.data.startswith("group_edit:"))
async def handle_group_edit(callback: CallbackQuery, state: FSMContext):
    """Start editing sheet name (which is also the group name)"""
    group_id = callback.data.split(":")[1]
    group = db.get_group(group_id)

    if not group:
        await safe_answer_callback(callback, "Group not found!", show_alert=True)
        return

    await safe_edit_message(
        callback,
        f"✏️ EDIT SHEET NAME\n\n"
        f"Current: {group['sheet_name']}\n\n"
        f"Enter new sheet name:\n"
        f"(Both group name and Google Sheets tab will be updated)"
    )

    await state.update_data(editing_group_id=group_id, old_sheet_name=group['sheet_name'])
    await state.set_state(GroupStates.waiting_for_edit_name)


@router.message(GroupStates.waiting_for_edit_name)
async def process_group_edit(message: Message, state: FSMContext):
    """Process sheet name edit - updates the Google Sheet tab and related records."""
    new_name = message.text.strip()

    if len(new_name) < 2 or len(new_name) > 50:
        await message.answer("❌ Name must be between 2 and 50 characters. Try again:")
        return

    data = await state.get_data()
    group_id = data.get('editing_group_id')
    old_sheet_name = data.get('old_sheet_name')

    # Rename Google Sheets tab first
    rename_success = sheets_manager.rename_sheet_tab(old_sheet_name, new_name)

    if not rename_success:
        await message.answer(
            f"❌ Failed to rename Google Sheets tab.\n"
            f"Please make sure the sheet '{old_sheet_name}' exists.",
            reply_markup=keyboards.get_teacher_keyboard()
        )
        await state.clear()
        return

    # ⭐ KEY FIX: Update all students' group_id to the new sheet name
    # This is the CRITICAL part - students must be updated
    students_updated = db.update_students_group_id(old_sheet_name, new_name)

    # Group info is sourced from the Google Sheet tab name.
    db.update_group(group_id, {
        'name': new_name,
        'sheet_name': new_name
    })

    # 🔄 Auto-refresh groups cache after rename
    teacher_id = str(message.from_user.id)
    db.get_teacher_groups(teacher_id, force_refresh=True)
    print(f"🔄 Auto-refreshed groups cache after rename")

    # Success!
    await message.answer(
        f"✅ Sheet renamed successfully!\n\n"
        f"Old: {old_sheet_name}\n"
        f"New: {new_name}\n\n"
        f"📊 Updated:\n"
        f"  • Google Sheets tab ✅\n"
        f"  • {students_updated} student(s) group_id ✅\n"
        f"  • Groups cache refreshed ✅\n\n"
        f"All done! 🎉",
        reply_markup=keyboards.get_teacher_keyboard()
    )

    await state.clear()


@router.callback_query(F.data == "groups:refresh")
async def handle_groups_refresh(callback: CallbackQuery):
    """Refresh groups from Google Sheets"""
    await safe_answer_callback(callback, "🔄 Refreshing groups...", show_alert=False)

    # Force refresh from Google Sheets
    teacher_id = str(callback.from_user.id)
    groups = db.get_teacher_groups(teacher_id, force_refresh=True)

    text = "👥 GROUP MANAGEMENT\n\n"
    if groups:
        text += f"✅ Refreshed! You have {len(groups)} group(s):\n"
        for group in groups:
            text += f"  • {group['name']} ({group['sheet_name']})\n"
    else:
        text += "You don't have any groups yet.\nCreate your first group to get started!"

    await safe_edit_message(
        callback,
        text,
        reply_markup=keyboards.get_groups_management_keyboard(teacher_id)
    )


@router.callback_query(F.data.startswith("group_delete:"))
async def handle_group_delete(callback: CallbackQuery):
    """Delete group (with confirmation)"""
    group_id = callback.data.split(":")[1]
    group = db.get_group(group_id)

    if not group:
        await safe_answer_callback(callback, "Group not found!", show_alert=True)
        return

    # Check if group has students
    students = db.get_all_users(role='student', status='active', group_id=group_id)

    if students:
        await safe_edit_message(
            callback,
            f"WARNING\n\n"
            f"Group '{group['name']}' has {len(students)} student(s).\n"
            f"Deleting this group would delete the entire Google Sheet tab and all rows in it.\n\n"
            f"Move or delete the students first, then try again.",
            reply_markup=keyboards.get_back_keyboard(f"group_view:{group_id}")
        )
        await safe_answer_callback(callback, "Group is not empty", show_alert=True)
        return
    else:
        await safe_edit_message(
            callback,
            f"DELETE GROUP\n\n"
            f"Are you sure you want to delete '{group['name']}'?",
            reply_markup=keyboards.get_confirmation_keyboard("delete_group_confirm", group_id)
        )


@router.callback_query(F.data.startswith("confirm:delete_group_confirm:"))
async def handle_group_delete_confirm(callback: CallbackQuery):
    """Confirm group deletion"""
    group_id = callback.data.split(":")[2]
    students = db.get_all_users(role='student', status='active', group_id=group_id)
    if students:
        await safe_edit_message(
            callback,
            "Cannot delete this group because it still has students.\n\n"
            "Move or delete the students first, then try again.",
            reply_markup=keyboards.get_back_keyboard(f"group_view:{group_id}")
        )
        await safe_answer_callback(callback, "Group is not empty", show_alert=True)
        return

    # Delete the group
    success = db.delete_group(group_id)

    if success:
        # 🔄 Auto-refresh groups cache after deletion
        teacher_id = str(callback.from_user.id)
        db.get_teacher_groups(teacher_id, force_refresh=True)
        print(f"🔄 Auto-refreshed groups cache after group deletion")

        await safe_edit_message(
            callback,
            "✅ Group deleted successfully!",
            reply_markup=keyboards.get_back_keyboard("settings:groups")
        )
    else:
        await safe_answer_callback(callback, "❌ Failed to delete group", show_alert=True)
